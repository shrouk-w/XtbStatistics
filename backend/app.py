from __future__ import annotations

import io
import math
import mimetypes
import os
import re
import time as time_module
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import httpx
import yfinance as yf
from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .database import get_db, init_db
from .models import CashOperation, ClosedPosition, OpenPosition, PendingOrder, LatestPrice

COMMENT_RE = re.compile(
    r"(?:OPEN|CLOSE) BUY (?P<quantity>\d+(?:\.\d+)?)(?:/(?P<total>\d+(?:\.\d+)?))? @ (?P<price>\d+(?:\.\d+)?)"
)
YAHOO_SUFFIX_MAP = {
    ".PL": ".WA",
    ".US": "",
    ".UK": ".L",
    ".FR": ".PA",
    ".NL": ".AS",
}
CURRENCY_SUFFIX_FALLBACK = {
    ".WA": "PLN",
    ".DE": "EUR",
    ".FR": "EUR",
    ".NL": "EUR",
    ".PA": "EUR",
    ".MI": "EUR",
    ".AS": "EUR",
    ".BR": "EUR",
    ".L": "GBP",
    ".LON": "GBP",
    ".UK": "GBP",
}
FX_SYMBOLS = {
    "EUR": "EURPLN=X",
    "USD": "USDPLN=X",
    "GBP": "GBPPLN=X",
    "PLN": None,
}
PROXY_ENV_NAMES = ["HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "http_proxy", "https_proxy", "all_proxy"]
BENCHMARKS = {
    "wig20": {"symbol": "WIG20.WA", "label": "WIG20"},
    "sp500": {"symbol": "^GSPC", "label": "S&P 500"},
    "nasdaq": {"symbol": "^IXIC", "label": "Nasdaq Composite"},
    "gold": {"symbol": "GC=F", "label": "Gold futures"},
    "bitcoin": {"symbol": "BTC-USD", "label": "Bitcoin"},
}

YAHOO_TO_STOOQ_MAP = {
    "CDR.WA": "CDR", "XTB.WA": "XTB", "PZU.WA": "PZU", "PXM.WA": "PXM", "TPE.WA": "TPE",
    "KGN.WA": "KGN", "INL.WA": "INL", "DIG.WA": "DIG", "ETL.WA": "ETL", "PAS.WA": "PAS",
    "11B.WA": "11B", "ACP.WA": "ACP", "ALE.WA": "ALE", "LPP.WA": "LPP", "PKN.WA": "PKN",
    "PKO.WA": "PKO", "PEO.WA": "PEO", "KGH.WA": "KGH", "DNP.WA": "DNP", "JSW.WA": "JSW",
    "^GSPC": "^SPX", "^IXIC": "^NDX", "BTC-USD": "BTCUSD", "GC=F": "XAUUSD",
}

mimetypes.add_type("application/javascript", ".js")
mimetypes.add_type("application/javascript", ".mjs")

@dataclass
class PositionEvent:
    ticker: str
    yahoo_ticker: str
    quantity_delta: float
    trade_date: date

@dataclass
class CashEvent:
    event_date: date
    amount: float
    operation_type: str

from pydantic import BaseModel

class ManualOperation(BaseModel):
    time: str # YYYY-MM-DD
    operation_type: str
    symbol: str | None = None
    amount: float
    comment: str | None = None

_portfolio_cache: dict[str, Any] = {"value": None, "expires_at": 0.0}
_PORTFOLIO_CACHE_TTL = 300  # seconds


def _invalidate_portfolio_cache() -> None:
    _portfolio_cache["value"] = None
    _portfolio_cache["expires_at"] = 0.0


def create_app() -> FastAPI:
    init_db()
    app = FastAPI(title="XTB Portfolio History")
    app.add_middleware(
        CORSMiddleware,
        allow_origin_regex=r"^https?://(localhost|127\.0\.0\.1)(:\d+)?$",
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    @app.get("/api/health")
    def health() -> dict[str, str]:
        return {"status": "ok"}

    @app.post("/api/portfolio/manual")
    async def add_manual_operation(op: ManualOperation, db: Session = Depends(get_db)):
        new_op = CashOperation(
            operation_type=op.operation_type,
            symbol=op.symbol,
            time=datetime.strptime(op.time, "%Y-%m-%d"),
            amount=op.amount,
            comment=op.comment
        )
        db.add(new_op)
        db.commit()
        _invalidate_portfolio_cache()
        return {"status": "success"}

    @app.get("/api/portfolio")
    async def get_portfolio(db: Session = Depends(get_db)) -> dict[str, Any]:
        # Serve from in-memory cache when fresh — yfinance + FX fetches are ~8s
        now = time_module.time()
        cached = _portfolio_cache["value"]
        if cached is not None and _portfolio_cache["expires_at"] > now:
            return cached
        # Fetch from DB
        db_cash = db.query(CashOperation).all()
        db_pending = db.query(PendingOrder).all()
        
        cash_events: list[CashEvent] = []
        external_cash_events: list[tuple[date, float]] = []
        trade_cash_events: list[tuple[date, float]] = []
        position_events: list[PositionEvent] = []

        for op in db_cash:
            event_date = op.time.date()
            cash_events.append(CashEvent(event_date=event_date, amount=op.amount, operation_type=op.operation_type))
            
            # Normalize operation types from DB
            op_type_raw = str(op.operation_type or "").strip().lower()
            
            if op_type_raw in {"stock purchase", "stock sale", "stock sell"}:
                trade_cash_events.append((event_date, op.amount))
            elif is_external_cash_operation(op_type_raw):
                external_cash_events.append((event_date, op.amount))
            
            # If it's a purchase/sell, we need quantity deltas
            if op_type_raw in {"stock purchase", "stock sale", "stock sell"}:
                quantity = extract_quantity(op.comment)
                if quantity:
                    delta = quantity if op_type_raw == "stock purchase" else -quantity
                    position_events.append(
                        PositionEvent(
                            ticker=op.symbol,
                            yahoo_ticker=to_yahoo_symbol(op.symbol),
                            quantity_delta=delta,
                            trade_date=event_date,
                        )
                    )

        # Include Pending Orders Margin as tied up cash (negative impact on available cash)
        for order in db_pending:
            if order.margin:
                event_date = order.open_time.date()
                # Assuming margin is positive in DB, it subtracts from available cash
                cash_events.append(CashEvent(event_date=event_date, amount=-float(order.margin), operation_type="Pending Order Margin"))

        if not cash_events and not position_events:
            print("DEBUG: No data found in database.")
            empty = {"summary": {}, "series": [], "holdings": [], "benchmarks": {}, "warnings": ["No data in database."]}
            _portfolio_cache["value"] = empty
            _portfolio_cache["expires_at"] = now + _PORTFOLIO_CACHE_TTL
            return empty

        portfolio = build_portfolio_history(cash_events, external_cash_events, trade_cash_events, position_events, db)
        _portfolio_cache["value"] = portfolio
        _portfolio_cache["expires_at"] = now + _PORTFOLIO_CACHE_TTL
        return portfolio

    @app.post("/api/portfolio/analyze")
    async def analyze_portfolio(
        files: list[UploadFile] = File(...),
        persist: bool = Form(False),
        db: Session = Depends(get_db)
    ) -> dict[str, Any]:
        if not files:
            raise HTTPException(status_code=400, detail="No files uploaded.")

        cash_events: list[CashEvent] = []
        external_cash_events: list[tuple[date, float]] = []
        trade_cash_events: list[tuple[date, float]] = []
        position_events: list[PositionEvent] = []
        warnings: list[str] = []

        for upload in files:
            filename = upload.filename or "uploaded.xlsx"
            if not filename.lower().endswith(".xlsx"):
                warnings.append(f"Skipped unsupported file: {filename}")
                continue

            payload = await upload.read()
            parsed = parse_xtb_workbook(payload, filename)
            cash_events.extend(parsed["cash_events"])
            external_cash_events.extend(parsed["external_cash_events"])
            trade_cash_events.extend(parsed["trade_cash_events"])
            position_events.extend(parsed["position_events"])
            warnings.extend(parsed["warnings"])

            if persist:
                workbook = load_workbook(io.BytesIO(payload), data_only=True)
                
                # Persistence for Cash Operations
                if "Cash Operations" in workbook.sheetnames:
                    sheet = workbook["Cash Operations"]
                    for row in sheet.iter_rows(min_row=6, values_only=True):
                        if not row or not row[0]: continue
                        op = CashOperation(
                            operation_type=str(row[0]).strip(),
                            symbol=(row[1] or "").strip(),
                            time=row[3] if isinstance(row[3], datetime) else None,
                            amount=float(row[4]) if row[4] is not None else 0.0,
                            comment=(row[6] or "").strip()
                        )
                        if op.time:
                            db.add(op)

                # Persistence for Pending Orders
                pending_sheet_name = next((s for s in workbook.sheetnames if "PENDING ORDERS" in s.upper()), None)
                if pending_sheet_name:
                    sheet = workbook[pending_sheet_name]
                    for row in sheet.iter_rows(min_row=9, values_only=True):
                        if not row or row[0] is None: continue
                        try:
                            order = PendingOrder(
                                order_id=row[0],
                                symbol=row[1],
                                margin=float(row[5]) if row[5] is not None else 0.0,
                                open_time=row[12] if isinstance(row[12], datetime) else None
                            )
                            if order.order_id:
                                db.merge(order)
                        except: continue

        if persist:
            try:
                db.commit()
                _invalidate_portfolio_cache()
            except Exception as e:
                db.rollback()
                warnings.append(f"Database sync failed: {str(e)}")

        if not cash_events and not position_events:
            raise HTTPException(status_code=400, detail="No XTB transactions found in uploaded files.")

        portfolio = build_portfolio_history(cash_events, external_cash_events, trade_cash_events, position_events, db)
        portfolio["warnings"].extend(warnings)
        return portfolio

    frontend_dist = Path(__file__).resolve().parent.parent / "frontend" / "dist"
    if frontend_dist.exists():
        app.mount("/assets", StaticFiles(directory=frontend_dist / "assets"), name="assets")
        @app.get("/")
        def index(): return FileResponse(frontend_dist / "index.html")
    return app

def parse_xtb_workbook(payload: bytes, filename: str) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(payload), data_only=True)
    ce = []; ece = []; tce = []; pe = []; ws = []
    if "Cash Operations" not in wb.sheetnames: ws.append(f"{filename}: missing 'Cash Operations' sheet."); return {"cash_events": ce, "external_cash_events": ece, "trade_cash_events": tce, "position_events": pe, "warnings": ws}
    for row in wb["Cash Operations"].iter_rows(min_row=6, values_only=True):
        if not row or not row[0]: continue
        ot = str(row[0]).strip(); xt = (row[1] or "").strip(); et = row[3]; am = row[4]; cm = (row[6] or "").strip()
        if not isinstance(et, datetime) or am is None: continue
        ed = et.date(); cam = float(am); ce.append(CashEvent(event_date=ed, amount=cam, operation_type=ot))
        if ot in {"Stock purchase", "Stock sale", "Stock sell"}: tce.append((ed, cam))
        elif is_external_cash_operation(ot): ece.append((ed, cam))
        if ot in {"Stock purchase", "Stock sale", "Stock sell"}:
            qty = extract_quantity(cm)
            if qty is None: ws.append(f"{filename}: could not parse quantity from '{cm}'."); continue
            pe.append(PositionEvent(ticker=xt, yahoo_ticker=to_yahoo_symbol(xt), quantity_delta=qty if ot == "Stock purchase" else -qty, trade_date=ed))
    psn = next((s for s in wb.sheetnames if "PENDING ORDERS" in s.upper()), None)
    if psn:
        for row in wb[psn].iter_rows(min_row=9, values_only=True):
            if not row or row[0] is None: continue
            try:
                m = float(row[5]) if row[5] is not None else 0.0; ot = row[12]
                if m > 0 and isinstance(ot, datetime): ce.append(CashEvent(event_date=ot.date(), amount=-m, operation_type="Pending Order Margin"))
            except: continue
    return {"cash_events": ce, "external_cash_events": ece, "trade_cash_events": tce, "position_events": pe, "warnings": ws}

def is_external_cash_operation(ot: str) -> bool:
    n = ot.lower()
    if any(t in n for t in ["stock purchase", "stock sale", "stock sell", "trade"]): return False
    if any(e in n for e in ["divident", "interest"]): return False
    return any(k in n for k in ["deposit", "withdrawal", "transfer", "wplata", "wyplata"])

def to_yahoo_symbol(xt: str) -> str:
    for s, r in YAHOO_SUFFIX_MAP.items():
        if xt.endswith(s): return xt[: -len(s)] + r
    return xt

def extract_quantity(cm: str) -> float | None:
    m = COMMENT_RE.search(cm)
    return float(m.group("quantity")) if m else None

def build_portfolio_history(ce, ece, tce, pe, db) -> dict[str, Any]:
    ad = [e.event_date for e in ce] + [e.trade_date for e in pe]
    if not ad: return {"summary": {}, "series": [], "holdings": [], "benchmarks": {}, "warnings": ["No dates found."]}
    sd = min(ad); ed = date.today(); idx = pd.date_range(start=sd, end=ed, freq="D")
    cs = pd.Series(0.0, index=idx)
    for e in ce: cs.loc[pd.Timestamp(e.event_date)] += float(e.amount)
    cs = cs.cumsum()
    ecs = pd.Series(0.0, index=idx)
    for d, a in ece: ecs.loc[pd.Timestamp(d)] += float(a)
    ecs = ecs.cumsum()
    if pe:
        qdf = build_quantities_frame(pe, idx)
        pdf, opdf, fcm, ws = build_price_frame(qdf, sd, ed, db)
        mv = qdf * pdf; hv = mv.sum(axis=1); ch = summarize_current_holdings(qdf, pdf, opdf, fcm)
    else: qdf = pd.DataFrame(index=idx); pdf = pd.DataFrame(index=idx); opdf = pd.DataFrame(index=idx); ws = []; hv = pd.Series(0.0, index=idx); ch = []
    tv = cs.add(hv, fill_value=0.0); pv = tv.subtract(ecs, fill_value=0.0)
    rdf = pd.DataFrame({"date": idx.strftime("%Y-%m-%d"), "cash": cs.round(2), "holdingsValue": hv.round(2), "totalValue": tv.round(2), "externalCashFlow": ecs.round(2), "profitValue": pv.round(2)})
    sm = {"startDate": rdf.iloc[0]["date"], "endDate": rdf.iloc[-1]["date"], "currentCash": round(float(cs.iloc[-1]), 2), "currentHoldingsValue": round(float(hv.iloc[-1]), 2), "currentTotalValue": round(float(tv.iloc[-1]), 2), "currentProfitValue": round(float(pv.iloc[-1]), 2), "netExternalCashFlow": round(float(ecs.iloc[-1]), 2), "currentProfitPercent": round((float(pv.iloc[-1]) / float(ecs.iloc[-1])) * 100, 2) if not math.isclose(float(ecs.iloc[-1]), 0.0, abs_tol=1e-9) else 0.0, "peakValue": round(float(tv.max()), 2), "lowestValue": round(float(tv.min()), 2)}
    bs, bws = build_benchmark_series(sd, ed, idx, tce, db)
    ws.extend(bws); return {"summary": sm, "series": rdf.to_dict(orient="records"), "holdings": ch, "benchmarks": bs, "warnings": ws}

def build_quantities_frame(pe, idx) -> pd.DataFrame:
    ts = sorted({e.yahoo_ticker for e in pe}); qs = pd.DataFrame(0.0, index=idx, columns=ts)
    for e in pe: qs.loc[pd.Timestamp(e.trade_date), e.yahoo_ticker] += e.quantity_delta
    return qs.cumsum()

from dotenv import load_dotenv
load_dotenv()

def get_yf_session(): return None
def get_investing_fx(c) -> float | None:
    if c == "PLN": return 1.0
    pm = {"USD": "usd-pln", "EUR": "eur-pln", "GBP": "gbp-pln", "CHF": "chf-pln"}; p = pm.get(c.upper())
    if not p: return None
    u = f"https://www.investing.com/currencies/{p}"; h = {'User-Agent': 'Mozilla/5.0'}
    try:
        r = httpx.get(u, headers=h, timeout=10.0)
        if r.status_code == 200:
            m = re.search(r'\"ask\":\s*\"([\d\.,]+)\"', r.text)
            if m: return float(m.group(1).replace(',', ''))
            m = re.search(r'instrument-price-last\">([\d\.,]+)', r.text)
            if m: return float(m.group(1).replace(',', ''))
    except: pass
    return None

def fetch_stooq_price(s) -> float | None:
    ak = os.getenv("STOOQ_API_KEY")
    if not ak: return None
    ss = YAHOO_TO_STOOQ_MAP.get(s, s.replace(".WA", "").replace(".PL", ""))
    u = f"https://stooq.pl/q/d/l/?s={ss.lower()}&i=d&apikey={ak}"
    try:
        r = httpx.get(u, headers={'User-Agent': 'Mozilla/5.0'}, timeout=10.0)
        if r.status_code == 200:
            if "Uzyskaj apikey" in r.text: return None
            df = pd.read_csv(io.StringIO(r.text))
            if not df.empty and 'Zamkniecie' in df.columns: return float(df['Zamkniecie'].iloc[-1])
    except: pass
    return None

def fetch_stooq_history(s) -> pd.Series:
    ak = os.getenv("STOOQ_API_KEY")
    if not ak: return pd.Series(dtype=float)
    ss = YAHOO_TO_STOOQ_MAP.get(s, s.replace(".WA", "").replace(".PL", ""))
    u = f"https://stooq.pl/q/d/l/?s={ss.lower()}&i=d&apikey={ak}"
    try:
        r = httpx.get(u, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15.0)
        if r.status_code == 200:
            if "Uzyskaj apikey" in r.text: return pd.Series(dtype=float)
            df = pd.read_csv(io.StringIO(r.text))
            date_col = 'Data' if 'Data' in df.columns else 'Date'
            close_col = 'Zamkniecie' if 'Zamkniecie' in df.columns else 'Close'
            if date_col in df.columns and close_col in df.columns:
                df[date_col] = pd.to_datetime(df[date_col]); df = df.set_index(date_col)
                return df[close_col].astype(float).sort_index()
    except: pass
    return pd.Series(dtype=float)

# Global to store API key validity for the current session
STOOQ_KEY_CHECKED = False
STOOQ_KEY_ERROR = None

def build_price_frame(qdf, sd, ed, db) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, str], list[str]]:
    idx = pd.date_range(start=sd, end=ed, freq="D"); ps = pd.DataFrame(index=idx); ops = pd.DataFrame(index=idx); cm = {}; ws = []
    ak = os.getenv("STOOQ_API_KEY")
    
    global STOOQ_KEY_CHECKED, STOOQ_KEY_ERROR
    
    if not ak: 
        ws.append("STOOQ_API_ERROR: Klucz API Stooq nie został skonfigurowany.")
    elif not STOOQ_KEY_CHECKED:
        STOOQ_KEY_CHECKED = True
        try:
            tr = httpx.get(f"https://stooq.pl/q/d/l/?s=wig20&i=d&apikey={ak}", headers={'User-Agent': 'Mozilla/5.0'}, timeout=7.0)
            if tr.status_code == 200 and "Uzyskaj apikey" in tr.text:
                STOOQ_KEY_ERROR = "STOOQ_API_ERROR: Twój klucz API Stooq wygasł lub jest nieprawidłowy. Sprawdź README.md, aby dowiedzieć się jak go zdobyć i odśwież aplikację."
            else:
                STOOQ_KEY_ERROR = None
        except:
            STOOQ_KEY_CHECKED = False # Retry next time if network failed
    
    if STOOQ_KEY_ERROR:
        ws.append(STOOQ_KEY_ERROR)
    
    uts = sorted(list(set(qdf.columns.tolist()) | {b["symbol"] for b in BENCHMARKS.values()}))
    for t in uts: cm[t] = resolve_currency(t)
    
    data = pd.DataFrame()
    try:
        print(f"DEBUG: Attempting Yahoo bulk download for {len(uts)} symbols...")
        data = yf.download(uts, start=sd.isoformat(), end=(ed + timedelta(days=1)).isoformat(), auto_adjust=False, progress=False, group_by="column", threads=True)
    except:
        print("DEBUG: Yahoo bulk download FAILED.")
    
    ucs = set(cm.values()) - {"PLN"}; frs = {}
    for c in ucs:
        r = get_investing_fx(c)
        if r:
            frs[c] = r; lp = db.query(LatestPrice).filter(LatestPrice.symbol == c).first()
            if not lp: lp = LatestPrice(symbol=c); db.add(lp)
            lp.price = r; lp.currency = "PLN"; lp.last_updated = datetime.utcnow()
        else:
            lp = db.query(LatestPrice).filter(LatestPrice.symbol == c).first()
            if lp: frs[c] = lp.price
            else: frs[c] = {"USD": 4.0, "EUR": 4.3, "GBP": 5.0}.get(c, 1.0); ws.append(f"FX Fallback for {c}")

    bts = [b["symbol"] for b in BENCHMARKS.values()]
    for bt in bts:
        h = pd.Series(dtype=float)
        if not data.empty:
            try:
                if len(uts) > 1:
                    if "Close" in data.columns: h = data["Close"][bt]
                    elif ("Close", bt) in data.columns: h = data[("Close", bt)]
                else:
                    if "Close" in data.columns: h = data["Close"]
            except: pass
        
        if h.empty or h.isnull().all() or bt == "WIG20.WA":
            print(f"DEBUG: Yahoo failed for benchmark {bt}. Trying Stooq history...")
            sh = fetch_stooq_history(bt)
            if not sh.empty:
                print(f"DEBUG: Stooq history SUCCESS for {bt} ({len(sh)} points)")
                h = sh
            else:
                print(f"DEBUG: Stooq history FAILED for {bt}. Trying single price...")
                sp = fetch_stooq_price(bt)
                if sp:
                    print(f"DEBUG: Stooq single price SUCCESS for {bt}: {sp}")
                    h = pd.Series(sp, index=idx)
        
        if not h.empty and not h.isnull().all():
            cv = float(h.iloc[-1])
            if cv > 0:
                lp = db.query(LatestPrice).filter(LatestPrice.symbol == bt).first()
                if not lp: lp = LatestPrice(symbol=bt); db.add(lp)
                lp.price = cv; lp.currency = "INDEX" if bt.startswith("^") or "WIG" in bt else "USD"; lp.last_updated = datetime.utcnow()

    xrm = {}
    for o in db.query(OpenPosition).all(): xrm[to_yahoo_symbol(o.symbol)] = o.symbol
    for c in db.query(ClosedPosition).all(): xrm[to_yahoo_symbol(c.symbol)] = c.symbol
    
    for t in uts:
        if t in bts: continue
        h = pd.Series(dtype=float)
        if not data.empty:
            try:
                if len(uts) > 1:
                    if "Close" in data.columns: h = data["Close"][t]
                    elif ("Close", t) in data.columns: h = data[("Close", t)]
                else:
                    if "Close" in data.columns: h = data["Close"]
            except: pass
        
        if h.empty or h.isnull().all():
            print(f"DEBUG: Yahoo failed for ticker {t}. Trying Stooq history...")
            sh = fetch_stooq_history(t)
            if not sh.empty:
                print(f"DEBUG: Stooq history SUCCESS for {t} ({len(sh)} points)")
                h = sh
            else:
                print(f"DEBUG: Stooq history FAILED for {t}. Trying single price...")
                sp = fetch_stooq_price(t)
                if sp:
                    print(f"DEBUG: Stooq single price SUCCESS for {t}: {sp}")
                    h = pd.Series(sp, index=idx)
        
        if not h.empty and not h.isnull().all():
            cv = float(h.iloc[-1])
            if cv > 0:
                lp = db.query(LatestPrice).filter(LatestPrice.symbol == t).first()
                if not lp: lp = LatestPrice(symbol=t); db.add(lp)
                lp.price = cv; lp.currency = cm.get(t); lp.last_updated = datetime.utcnow()
        
        if h.empty or h.isnull().all():
            lp = db.query(LatestPrice).filter(LatestPrice.symbol == t).first()
            if lp and lp.price: h = pd.Series(lp.price, index=idx)
            else:
                xs = xrm.get(t, t.replace(".WA", ".PL")); lop = db.query(OpenPosition).filter(OpenPosition.symbol == xs).first(); mp = lop.market_price if lop else 0.0
                lcp = db.query(ClosedPosition).filter(ClosedPosition.symbol == xs).order_by(ClosedPosition.close_time.desc()).first(); cp = lcp.close_price if lcp else 0.0
                if mp > 0: h = pd.Series(mp, index=idx)
                elif cp > 0: h = pd.Series(cp, index=idx)
                else: h = pd.Series(0.0, index=idx)
        
        h = h.reindex(idx).ffill().bfill().fillna(0.0); ops[t] = h.round(6); c = cm[t]
        if c != "PLN": ps[t] = (h * frs.get(c, 1.0)).round(6)
        else: ps[t] = h.round(6)
    
    try: db.commit()
    except: db.rollback()
    return ps.fillna(0.0), ops.fillna(0.0), cm, ws

def download_close_series(s, sd, ed) -> pd.Series:
    with without_dead_local_proxy():
        try: data = yf.download(s, start=sd.isoformat(), end=(ed + timedelta(days=1)).isoformat(), auto_adjust=False, progress=False, threads=False)
        except: return pd.Series(dtype=float)
    if data is None or data.empty: return pd.Series(dtype=float)
    if isinstance(data.columns, pd.MultiIndex): cs = data["Close"].iloc[:, 0]
    else: cs = data["Close"]
    cs.index = pd.to_datetime(cs.index).tz_localize(None); return cs.astype(float).sort_index()

def resolve_currency(s) -> str:
    g = guess_currency_from_symbol(s)
    if g != "USD": return g
    with without_dead_local_proxy():
        t = yf.Ticker(s)
        try:
            fi = t.fast_info
            if fi:
                curr = fi.get("currency")
                if curr: return str(curr).upper()
        except: pass
    return g

def guess_currency_from_symbol(s) -> str:
    us = s.upper()
    for sx, c in CURRENCY_SUFFIX_FALLBACK.items():
        if us.endswith(sx): return c
    return "USD"

def summarize_current_holdings(qdf, pdf, opdf, cm) -> list[dict[str, Any]]:
    hs = []; lq = qdf.iloc[-1]; lps = pdf.iloc[-1] if not pdf.empty else pd.Series(dtype=float); lops = opdf.iloc[-1] if not opdf.empty else pd.Series(dtype=float)
    for t, q in lq.items():
        if math.isclose(float(q), 0.0, abs_tol=1e-9): continue
        p = float(lps.get(t, 0.0)); op = float(lops.get(t, 0.0)); oc = cm.get(t, "PLN")
        hs.append({"ticker": t, "quantity": round(float(q), 6), "pricePln": round(p, 2), "marketValuePln": round(p * float(q), 2), "priceOriginal": round(op, 4), "marketValueOriginal": round(op * float(q), 2), "originalCurrency": oc, "currency": oc, "insight": fetch_holding_insight(t)})
    hs.sort(key=lambda x: x["marketValuePln"], reverse=True); return hs

def build_benchmark_series(sd, ed, idx, tce, db) -> tuple[dict[str, Any], list[str]]:
    bs = {}; ws = []; tcb = {}
    for d, a in tce: tcb[d] = tcb.get(d, 0.0) + float(a)
    for k, cfg in BENCHMARKS.items():
        s = cfg["symbol"]; h = pd.Series(dtype=float)
        print(f"DEBUG: Building benchmark {k} ({s})...")
        h = fetch_stooq_history(s)
        if h.empty:
            print(f"DEBUG: Stooq history empty for benchmark {s}. Trying Yahoo...")
            h = download_close_series(s, sd, ed)
        else:
            print(f"DEBUG: Stooq history SUCCESS for benchmark {s} ({len(h)} points)")
            
        if h.empty:
            print(f"DEBUG: Yahoo failed for benchmark {s}. Trying LatestPrice...")
            lp = db.query(LatestPrice).filter(LatestPrice.symbol == s).first()
            if lp and lp.price: h = pd.Series(lp.price, index=idx)
            else: ws.append(f"No benchmark history for {cfg['label']}"); continue
        
        h = h.reindex(idx).ffill().bfill()
        print(f"DEBUG: Final benchmark series for {s} has {len(h)} points and last value {h.iloc[-1] if not h.empty else 'N/A'}")
        
        fp = float(h[h > 0].iloc[0]) if not h[h > 0].empty else 0.0
        if math.isclose(fp, 0.0, abs_tol=1e-12): ws.append(f"Benchmark {cfg['label']} has no usable starting price."); continue
        sh = 0.0; ni = 0.0; vs = []; ivs = []; pvs = []
        for ts, p in h.items():
            ea = tcb.get(ts.date(), 0.0); p = float(p)
            if ea < 0: ia = -ea; sh += ia / p; ni += ia
            elif ea > 0: ss = min(sh, ea / p); sh -= ss; ni -= min(ni, ea)
            v = sh * p; vs.append(v); ivs.append(ni); pvs.append(v - ni)
        fv = next((v for v in vs if not math.isclose(v, 0.0, abs_tol=1e-9)), 0.0)
        rp = [((v / fv) - 1.0) * 100 if fv else 0.0 for v in vs]
        pts = pd.DataFrame({"date": idx.strftime("%Y-%m-%d"), "value": pd.Series(vs, index=idx).round(2), "netInvested": pd.Series(ivs, index=idx).round(2), "profitValue": pd.Series(pvs, index=idx).round(2), "returnPercent": pd.Series(rp, index=idx).round(2)})
        bs[k] = {"label": cfg["label"], "symbol": s, "series": pts.to_dict(orient="records")}
    return bs, ws

def fetch_holding_insight(s): return {"name": s, "sector": None, "recommendation": None, "recommendationMean": None, "analystCount": None, "targetMeanPrice": None, "targetCurrency": None, "source": "Yahoo Finance (Disabled)", "asOf": None, "summary": "Analiza wyłączona czasowo."}

@contextmanager
def without_dead_local_proxy():
    r = {}
    for e in PROXY_ENV_NAMES:
        v = os.environ.get(e)
        if v and "127.0.0.1:9" in v: r[e] = v; os.environ.pop(e, None)
    try: yield
    finally:
        for e, v in r.items(): os.environ[e] = v

app = create_app()
